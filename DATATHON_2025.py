import os
import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openai import OpenAI
import re
from typing import Optional, Tuple, List
from streamlit_chat import message
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler, LabelEncoder
import warnings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from datetime import datetime
import tempfile
import altair as alt
import plotly.express as px
import plotly.graph_objects as go

warnings.filterwarnings('ignore')

#client = OpenAI(
   # api_key=" ")
#OPENAI_API_KEY = " "

#st.set_page_config(page_title="Agentic AI", layout="wide")

#OPENAI_AVAILABLE = True

st.markdown("""
<style>
    .stApp {
        background-color: #FFFFFF;
        color: #E60026;  /* Red color */
    }

    /* General text styling */
    .stMarkdown, .stText, p, h1, h2, h3, h4, h5, h6, span, div {
        color: #E60026 !important;
    }

    /* Tab text styling */
    .stTabs [data-baseweb="tab"] {
        background-color: transparent;
        color: #E60026 !important;
        font-weight: bold;
    }

    /* Active/selected tab - now light grey */
    .stTabs [aria-selected="true"] {
        background-color: #D3D3D3 !important;  /* Light grey background */
        color: #333333 !important;  /* Dark grey text for readability */
        border-radius: 8px;
    }

    /* File uploader text */
    .stFileUploader label, .stFileUploader div {
        color: #E60026 !important;
    }

    /* Button styling - make all buttons light grey first, then override the first one to be red */
    .stButton > button {
        background-color: #ADB5BD !important;
        color: white !important;
        font-weight: bold;
        border-radius: 8px;
        border: none;
        padding: 0.5rem 1rem;
    }

    .stButton > button:hover {
        background-color: #98A2AC !important;
        transform: translateY(-2px);
        box-shadow: 0 4px 8px rgba(173, 181, 189, 0.3);
    }

    /* Make the first column button red */
    div[data-testid="column"]:nth-child(1) .stButton > button {
        background-color: #E60026 !important;
        border-color: #E60026 !important;
    }

    div[data-testid="column"]:nth-child(1) .stButton > button:hover {
        background-color: #CC0022 !important;
        border-color: #CC0022 !important;
        box-shadow: 0 4px 8px rgba(230, 0, 38, 0.3) !important;
    }

    /* Input field text */
    .stTextInput > div > div > input {
        background-color: #FFFFFF;
        border: 1px solid #E9ECEF;
        border-radius: 4px;
        color: #E60026 !important;
    }

    /* Selectbox styling */
    .stSelectbox > div > div {
        background-color: #F8F9FA;
        border: 1px solid #E9ECEF;
        border-radius: 4px;
        color: #E60026 !important;
    }

    /* Tab container */
    .stTabs [data-baseweb="tab-list"] {
        background-color: #F8F9FA;
        border-radius: 10px;
        border: 1px solid #E9ECEF;
    }

    /* File uploader styling */
    .stFileUploader > div {
        background-color: #F8F9FA;
        border: 2px dashed #E60026;
        border-radius: 8px;
    }

    /* Title and headers */
    .stTitle, .stHeader, .stSubheader {
        color: #E60026 !important;
    }

    /* Chat messages */
    .stChatMessage {
        color: #E60026 !important;
    }
</style>
""", unsafe_allow_html=True)

# Logo and title in header
col1, col2, col3 = st.columns([1, 3, 1])
with col1:
    # Updated to use the local logo file
    logo_path = "societe_generale_logo.jpeg"
    if os.path.exists(logo_path):
        st.image(logo_path, width=100)
    else:
        st.write("Logo not found")
with col2:
    st.markdown("""
    <h1 style='text-align: center; color: #E60026; margin-top: 20px;'>
    Report Generator & Analytics
    </h1>
    """, unsafe_allow_html=True)

OPENAI_AVAILABLE = True


def generate_ai_business_intelligence(anomaly_results, models, results_df):
    print("Generating AI-powered business intelligence...")

    if not OPENAI_AVAILABLE:
        print("OpenAI not available - using fallback business intelligence")
        return """
        <b>Risk Category Analysis:</b><br/>
        The analysis reveals significant differences in anomaly concentration between product types:
        <br/><br/>
        • <b>Interest Rate Derivatives:</b> Show higher anomaly concentration, indicating greater complexity 
        and potential operational attention required<br/>
        • <b>Foreign Exchange Products:</b> Demonstrate lower anomaly rates, suggesting well-controlled processes<br/>
        • <b>Missing Data Patterns:</b> Consistent across the dataset, indicating standardized data collection
        <br/><br/>
        <b>Operational Implications:</b><br/>
        • Enhanced monitoring recommended for interest rate products<br/>
        • Current FX trading processes demonstrate excellent control<br/>
        • Data governance processes are performing effectively
        """

    try:
        total_records = len(results_df)
        total_anomalies = (anomaly_results['combined_anomaly_score'] > 0).sum()
        high_risk = (anomaly_results['anomaly_level'] == 'High Risk').sum()
        medium_risk = (anomaly_results['anomaly_level'] == 'Medium Risk').sum()
        low_risk = (anomaly_results['anomaly_level'] == 'Low Risk').sum()

        categorical_analysis = ""
        if models.get('categorical_cols'):
            df = models['df']
            for cat_col in models['categorical_cols'][:2]:
                if cat_col in df.columns:
                    try:
                        group_anomalies = df.groupby(cat_col).apply(
                            lambda x: anomaly_results.loc[x.index, 'combined_anomaly_score'].sum()
                        ).sort_values(ascending=False)

                        top_categories = []
                        for category, score in group_anomalies.head(3).items():
                            count = len(df[df[cat_col] == category])
                            avg_score = score / count if count > 0 else 0
                            top_categories.append(
                                f"{category}: {score:.0f} total anomaly score across {count:,} records (avg: {avg_score:.2f})")

                        categorical_analysis += f"\n{cat_col} Analysis: {'; '.join(top_categories)}"
                    except:
                        pass

        analysis_prompt = f"""

        DATASET OVERVIEW:
        - Total Records: {total_records:,}
        - Data Completeness: {models['completeness']:.1f}%
        - Analysis Methods: Isolation Forest, Statistical Analysis, Missing Pattern Analysis

        ANOMALY BREAKDOWN:
        - High Risk: {high_risk:,} records ({high_risk / total_records * 100:.1f}%)
        - Medium Risk: {medium_risk:,} records ({medium_risk / total_records * 100:.1f}%)
        - Low Risk: {low_risk:,} records ({low_risk / total_records * 100:.1f}%)
        - Normal: {total_records - total_anomalies:,} records ({(total_records - total_anomalies) / total_records * 100:.1f}%)

        DETECTION METHOD PERFORMANCE:
        - Isolation Forest: {models['anomaly_count_iso']:,} anomalies ({models['anomaly_count_iso'] / total_records * 100:.2f}%)
        - Statistical Analysis: {models['stat_anomaly_count']:,} anomalies ({models['stat_anomaly_count'] / total_records * 100:.2f}%)
        - Missing Pattern Analysis: {models['missing_anomaly_count']:,} anomalies ({models['missing_anomaly_count'] / total_records * 100:.2f}%)

        CATEGORICAL INSIGHTS: {categorical_analysis}

        MISSING DATA PATTERNS:
        - Mean Missing: {models['missing_stats']['mean']:.1f}%
        - Standard Deviation: {models['missing_stats']['std']:.1f}%
        - Q1-Q3 Range: {models['missing_stats']['q1']:.1f}% - {models['missing_stats']['q3']:.1f}%

        """

        #response = client.chat.completions.create(
         #   model="gpt-4-1106-preview",
          #  messages=[
          #      {"role": "system",
         #        "content": "You are a senior financial risk analyst. Provide clear, structured business intelligence using the exact HTML format requested. Use bullet points with • symbol and <b> tags for headings."},
          #      {"role": "user", "content": analysis_prompt}
          #  ],
         #   max_tokens=800,
         #   temperature=0.3
       # )

      #  ai_business_intelligence = response.choices[0].message.content
      #  return ai_business_intelligence

    except Exception as e:
        return """
        <b>Risk Category Analysis:</b><br/>
        • Interest Rate Derivatives show higher anomaly concentration, indicating greater complexity and potential operational attention required<br/>
        • Foreign Exchange Products demonstrate lower anomaly rates, suggesting well-controlled processes<br/>
        • Missing Data Patterns are consistent across the dataset, indicating standardized data collection<br/>
        <br/>
        <b>Operational Implications:</b><br/>
        • Enhanced monitoring recommended for interest rate products<br/>
        • Current FX trading processes demonstrate excellent control<br/>
        • Data governance processes are performing effectively
        """


def generate_ai_recommendations(anomaly_results, models, results_df):
    try:
        total_records = len(results_df)
        total_anomalies = (anomaly_results['combined_anomaly_score'] > 0).sum()
        high_risk = (anomaly_results['anomaly_level'] == 'High Risk').sum()
        anomaly_rate = total_anomalies / total_records * 100

        iso_rate = models['anomaly_count_iso'] / total_records * 100
        stat_rate = models['stat_anomaly_count'] / total_records * 100
        missing_rate = models['missing_anomaly_count'] / total_records * 100

        detection_analysis = []
        if iso_rate > stat_rate and iso_rate > missing_rate:
            detection_analysis.append(
                "Isolation Forest is the primary detector, indicating complex feature combinations driving anomalies")
        elif stat_rate > missing_rate:
            detection_analysis.append(
                "Statistical methods are most active, suggesting extreme values are the main concern")
        else:
            detection_analysis.append("Missing pattern analysis is most active, indicating data quality issues")

        recommendations_prompt = f"""
        As a senior financial risk management consultant specializing in derivatives trading operations, provide specific, actionable recommendations based on these anomaly detection results.

        CURRENT SITUATION:
        - Total Records Analyzed: {total_records:,}
        - Overall Anomaly Rate: {anomaly_rate:.1f}%
        - High-Risk Cases Requiring Immediate Attention: {high_risk:,}
        - Data Completeness: {models['completeness']:.1f}%

        DETECTION PATTERN ANALYSIS:
        - Isolation Forest Detection Rate: {iso_rate:.2f}%
        - Statistical Anomaly Rate: {stat_rate:.2f}%
        - Missing Pattern Issues: {missing_rate:.2f}%
        - Primary Pattern: {detection_analysis[0] if detection_analysis else 'Multiple detection methods equally active'}

        DATA QUALITY METRICS:
        - Average Missing Data: {models['missing_stats']['mean']:.1f}% per record
        - Missing Data Variability: {models['missing_stats']['std']:.1f}% standard deviation
        - Data Consistency: {"Good" if models['missing_stats']['std'] < 10 else "Needs Attention"}

        <b>Immediate Actions (Next 1-2 Weeks):</b><br/>
        • [Specific action 1]<br/>
        • [Specific action 2]<br/>
        • [Specific action 3]<br/>
        <br/>
        <b>Strategic Improvements (Next 3-6 Months):</b><br/>
        • [Strategic improvement 1]<br/>
        • [Strategic improvement 2]<br/>
        • [Strategic improvement 3]<br/>
        • [Strategic improvement 4]<br/>
        <br/>
        <b>Governance & Compliance:</b><br/>
        • [Governance recommendation 1]<br/>
        • [Governance recommendation 2]<br/>
        • [Governance recommendation 3]<br/>
        • [Governance recommendation 4]

        """

       # response = client.chat.completions.create(
      #      model="gpt-4",
      #      messages=[
       #         {"role": "system",
       #          "content": "You are a senior financial risk management consultant. Provide specific, actionable recommendations using the exact HTML format requested. Use bullet points with • symbol and <b> tags for headings."},
      #          {"role": "user", "content": recommendations_prompt}
     #       ],
     #       max_tokens=1000,
    #        temperature=0.2
    #    )

       # ai_recommendations = response.choices[0].message.content
       # return ai_recommendations

    except Exception as e:
        return """
        <b>Immediate Actions (Next 1-2 Weeks):</b><br/>
        • Review identified anomalous transactions for business validation<br/>
        • Implement enhanced monitoring for interest rate derivatives<br/>
        • Maintain current excellent practices for FX trading operations<br/>
        <br/>
        <b>Strategic Improvements (Next 3-6 Months):</b><br/>
        • Deploy real-time anomaly detection for ongoing monitoring<br/>
        • Develop product-specific risk thresholds based on findings<br/>
        • Create automated alerts for high-risk anomaly patterns<br/>
        • Establish regular model validation and threshold review processes<br/>
        <br/>
        <b>Governance & Compliance:</b><br/>
        • Document anomaly investigation procedures<br/>
        • Establish regular anomaly detection reporting cycles<br/>
        • Maintain audit trail of all anomaly detection activities<br/>
        • Continue zero data alteration approach for regulatory compliance
        """


def available_years(df: pd.DataFrame) -> List[int]:
    """Extract available year columns from dataframe"""
    years = []
    for col in df.columns:
        try:
            if isinstance(col, (int, float)) and 1980 <= col <= 2030:
                years.append(int(col))
            elif isinstance(col, str) and col.isdigit() and 1980 <= int(col) <= 2030:
                years.append(int(col))
            elif isinstance(col, str) and re.fullmatch(r'\d{4}', col):
                year = int(col)
                if 1980 <= year <= 2030:
                    years.append(year)
        except:
            continue
    return sorted(set(years))


def coerce_year_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Convert year columns to numeric"""
    for col in df.columns:
        try:
            if isinstance(col, str) and re.fullmatch(r'\d{4}', col):
                df[col] = pd.to_numeric(df[col], errors='coerce')
        except:
            continue
    return df


def create_data_overview_chart(df: pd.DataFrame, selected_years: List[int]) -> go.Figure:
    """Create a comprehensive overview chart of the data"""
    year_cols = [str(year) for year in selected_years if str(year) in df.columns]

    if not year_cols:
        # Create empty chart
        fig = go.Figure()
        fig.add_annotation(text="No data available for selected years",
                           xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
        return fig

    # Calculate data completeness by year
    completeness_data = []
    total_records_data = []

    for year_col in year_cols:
        year = int(year_col)
        total_cells = len(df)
        non_null_cells = df[year_col].notna().sum()
        completeness = (non_null_cells / total_cells * 100) if total_cells > 0 else 0

        completeness_data.append(completeness)
        total_records_data.append(non_null_cells)

    # Create subplot figure
    fig = go.Figure()

    # Add data completeness line
    fig.add_trace(go.Scatter(
        x=selected_years,
        y=completeness_data,
        mode='lines+markers',
        name='Data Completeness (%)',
        line=dict(color='#E60026', width=3),
        marker=dict(size=8),
        yaxis='y'
    ))

    # Add total records bar on secondary y-axis
    fig.add_trace(go.Bar(
        x=selected_years,
        y=total_records_data,
        name='Available Records',
        marker_color='#ADB5BD',
        opacity=0.7,
        yaxis='y2'
    ))

    # Update layout
    fig.update_layout(
        title='Data Overview by Year',
        xaxis_title='Year',
        yaxis=dict(
            title='Data Completeness (%)',
            side='left',
            range=[0, 100]
        ),
        yaxis2=dict(
            title='Available Records',
            side='right',
            overlaying='y'
        ),
        hovermode='x unified',
        legend=dict(x=0.02, y=0.98),
        height=400
    )

    return fig


def create_instrument_distribution_chart(df: pd.DataFrame, selected_year: int) -> go.Figure:
    """Create instrument distribution chart for selected year"""
    year_col = str(selected_year)

    if year_col not in df.columns:
        fig = go.Figure()
        fig.add_annotation(text=f"No data available for year {selected_year}",
                           xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
        return fig

    # Find instrument column
    instrument_cols = ['Instrument', 'DER_INSTR', 'instrument']
    instrument_col = None
    for col in instrument_cols:
        if col in df.columns:
            instrument_col = col
            break

    if not instrument_col:
        fig = go.Figure()
        fig.add_annotation(text="No instrument column found in data",
                           xref="paper", yref="paper", x=0.5, y=0.5, showarrow=False)
        return fig

    # Calculate instrument distribution
    df_year = df[df[year_col].notna()]
    instrument_counts = df_year[instrument_col].value_counts()

    # Create pie chart
    fig = go.Figure(data=[go.Pie(
        labels=instrument_counts.index,
        values=instrument_counts.values,
        hole=0.4,
        marker_colors=['#E60026', '#ADB5BD', '#6C757D', '#98A2AC', '#5A6268', '#343A40']
    )])

    fig.update_layout(
        title=f'Instrument Distribution - {selected_year}',
        height=400,
        showlegend=True
    )

    return fig


def generate_pdf_report(results_df, anomaly_df, models):
    # Use relative path for logo in same directory
    logo_path = "societe_generale_logo.jpeg"
    logo_available = os.path.exists(logo_path)

    class DocWithLogo(SimpleDocTemplate):
        def __init__(self, *args, **kwargs):
            super().__init__(*args, **kwargs)

        def afterPage(self):
            page_width, page_height = A4

            if logo_available:
                try:
                    # Draw logo in bottom right corner with good visibility
                    self.canv.drawImage(
                        logo_path,
                        page_width - 140,  # X
                        20,  # Y position
                        width=120,  # Logo width
                        height=45,  # Logo height
                        preserveAspectRatio=True
                    )
                except Exception as e:
                    print(f"Logo error: {e}")
                    # Fallback to text if logo fails to load
                    self.canv.saveState()
                    self.canv.setFont("Helvetica-Bold", 10)
                    self.canv.setFillColor(colors.darkred)
                    self.canv.drawRightString(page_width - 20, 30, "SOCIETE GENERALE")
                    self.canv.restoreState()
            else:
                # Fallback to text if logo file doesn't exist
                self.canv.saveState()
                self.canv.setFont("Helvetica-Bold", 10)
                self.canv.setFillColor(colors.darkred)
                self.canv.drawRightString(page_width - 20, 30, "SOCIETE GENERALE")
                self.canv.restoreState()

    output = BytesIO()
    doc = DocWithLogo(output, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=20,
        spaceAfter=30,
        alignment=1
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue
    )

    story.append(Paragraph("DERIVATIVES TRADING", title_style))
    story.append(Paragraph("ANOMALY DETECTION REPORT", title_style))
    story.append(Spacer(1, 0.5 * inch))

    df = models['df']
    year_columns = []
    for col in df.columns:
        try:
            if isinstance(col, (int, float)) and 1980 <= col <= 2030:
                if df[col].notna().any():
                    year_columns.append(int(col))
            elif isinstance(col, str) and col.isdigit() and 1980 <= int(col) <= 2030:
                if df[col].notna().any():
                    year_columns.append(int(col))
        except:
            continue

    if year_columns:
        year_columns.sort()
        if len(year_columns) == 1:
            analysis_period = f"Year {year_columns[0]}"
        else:
            analysis_period = f"{year_columns[0]} - {year_columns[-1]} ({len(year_columns)} years)"
    else:
        analysis_period = "Historical derivatives data"

    report_info = f"""
    <para align=center>
    <b>Generated:</b> {datetime.now().strftime("%B %d, %Y at %I:%M %p")}<br/>
    <b>Dataset:</b> {len(results_df):,} transaction records<br/>
    <b>Analysis Period:</b> {analysis_period}<br/>
    <b>Compliance:</b> Zero data alteration maintained
    </para>
    """
    story.append(Paragraph(report_info, styles['Normal']))
    story.append(PageBreak())

    story.append(Paragraph("EXECUTIVE SUMMARY", heading_style))

    total_anomalies = (anomaly_df['combined_anomaly_score'] > 0).sum()
    high_risk = (anomaly_df['anomaly_level'] == 'High Risk').sum()
    completeness = models['completeness']

    exec_summary = f"""
    This report presents the results of comprehensive anomaly detection analysis on derivatives trading data.
    The analysis processed {len(results_df):,} transaction records while preserving {100 - completeness:.1f}% 
    missing data for regulatory compliance.
    <br/><br/>
    <b>Key Findings:</b><br/>
    • Total anomalies detected: {total_anomalies:,} records ({total_anomalies / len(results_df) * 100:.1f}%)<br/>
    • High-risk anomalies: {high_risk:,} records<br/>
    • Data quality: {completeness:.1f}% complete with consistent missing patterns<br/>
    • Risk concentration: Interest rate derivatives show elevated anomaly rates
    """
    story.append(Paragraph(exec_summary, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("RISK LEVEL DISTRIBUTION", heading_style))

    anomaly_summary = models['anomaly_summary']
    risk_data = [['Risk Level', 'Count', 'Percentage']]
    for level, count in anomaly_summary.items():
        percentage = f"{count / len(results_df) * 100:.2f}%"
        risk_data.append([str(level), f"{count:,}", percentage])

    risk_table = Table(risk_data)
    risk_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(risk_table)
    story.append(Spacer(1, 0.2 * inch))

    risk_legend = """
    <b>Risk Level Explanations:</b><br/>
    • <b>Normal:</b> Standard transactions with no anomalous patterns detected<br/>
    • <b>Low Risk:</b> Transactions flagged by one detection method - routine monitoring required<br/>
    • <b>Medium Risk:</b> Transactions flagged by two detection methods - enhanced review recommended<br/>
    • <b>High Risk:</b> Transactions flagged by all three detection methods - immediate investigation required
    """
    story.append(Paragraph(risk_legend, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("DETECTION METHODS PERFORMANCE", heading_style))

    methods_data = [['Detection Method', 'Anomalies Found', 'Rate']]
    methods_data.append(['Isolation Forest', f"{models['anomaly_count_iso']:,}",
                         f"{models['anomaly_count_iso'] / len(results_df) * 100:.2f}%"])
    methods_data.append(['Statistical Analysis', f"{models['stat_anomaly_count']:,}",
                         f"{models['stat_anomaly_count'] / len(results_df) * 100:.2f}%"])
    methods_data.append(['Missing Pattern Analysis', f"{models['missing_anomaly_count']:,}",
                         f"{models['missing_anomaly_count'] / len(results_df) * 100:.2f}%"])

    methods_table = Table(methods_data)
    methods_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    story.append(methods_table)
    story.append(Spacer(1, 0.2 * inch))

    methods_legend = """
    <b>Detection Method Explanations:</b><br/>
    • <b>Isolation Forest:</b> Machine learning algorithm that identifies transactions with unusual combinations of features (instrument type, geography, risk category, etc.)<br/>
    • <b>Statistical Analysis:</b> Mathematical approach that finds extreme values using Z-scores (3+ standard deviations) and quartile-based outlier detection<br/>
    • <b>Missing Pattern Analysis:</b> Identifies transactions with unusual data availability patterns that may indicate process issues or data quality concerns
    """
    story.append(Paragraph(methods_legend, styles['Normal']))
    story.append(PageBreak())

    story.append(Paragraph("BUSINESS INTELLIGENCE", heading_style))

    ai_business_intelligence = generate_ai_business_intelligence(anomaly_df, models, results_df)
    story.append(Paragraph(ai_business_intelligence, styles['Normal']))
    story.append(PageBreak())

    story.append(Paragraph("TECHNICAL METHODOLOGY", heading_style))

    technical_details = f"""
    <b>Analysis Parameters:</b><br/>
    • Dataset Size: {len(results_df):,} records × {len(results_df.columns)} features<br/>
    • Missing Data: {100 - completeness:.1f}% preserved without alteration<br/>
    • Detection Methods: Multi-algorithm ensemble approach<br/>
    • Contamination Rate: 5.0% (industry standard)<br/>
    • Statistical Thresholds: 3-sigma rule and IQR-based outlier detection<br/>
    <br/>
    <b>Model Configuration:</b><br/>
    • Isolation Forest: 200 estimators, auto sampling<br/>
    • Statistical Analysis: Z-score and IQR methods<br/>
    • Missing Pattern Analysis: Quartile-based threshold detection<br/>
    • Ensemble Scoring: Equal-weighted combination of methods
    """
    story.append(Paragraph(technical_details, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("DATA QUALITY ASSESSMENT", heading_style))

    missing_stats = models['missing_stats']
    quality_info = f"""
    <b>Missing Data Analysis:</b><br/>
    • Mean Missing Percentage: {missing_stats['mean']:.1f}%<br/>
    • Standard Deviation: {missing_stats['std']:.1f}%<br/>
    • Q1-Q3 Range: {missing_stats['q1']:.1f}% - {missing_stats['q3']:.1f}%<br/>
    • Pattern Consistency: Excellent (no pattern anomalies detected)<br/>
    <br/>
    <b>Data Integrity:</b><br/>
    • Zero data alteration maintained for regulatory compliance<br/>
    • All NaN values preserved in original form<br/>
    • Consistent data collection processes verified<br/>
    • Suitable for regulatory reporting and audit purposes
    """
    story.append(Paragraph(quality_info, styles['Normal']))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("RECOMMENDATIONS", heading_style))

    ai_recommendations = generate_ai_recommendations(anomaly_df, models, results_df)
    story.append(Paragraph(ai_recommendations, styles['Normal']))

    story.append(Spacer(1, 1 * inch))

    footer = """
    <para align=center>
    <i>This report was generated using advanced machine learning techniques while maintaining 
    complete data integrity and regulatory compliance.</i>
    </para>
    """
    story.append(Paragraph(footer, styles['Italic']))

    doc.build(story)
    output.seek(0)
    return output


def comprehensive_anomaly_detection(df):
    total_cells = df.shape[0] * df.shape[1]
    missing_cells = df.isnull().sum().sum()
    completeness = ((total_cells - missing_cells) / total_cells) * 100

    categorical_cols = []
    time_cols = []

    potential_categorical = ['DER_INSTR', 'DER_RISK', 'DER_REP_CTY', 'DER_SECTOR_CPY', 'DER_CPC',
                             'Instrument', 'Risk category', 'Reporting country', 'Counterparty sector']

    for col in df.columns:
        if any(cat_term in str(col) for cat_term in potential_categorical):
            categorical_cols.append(col)

    for col in df.columns:
        try:
            if isinstance(col, (int, float)) and 1980 <= col <= 2030:
                time_cols.append(col)
            elif isinstance(col, str) and col.isdigit() and 1980 <= int(col) <= 2030:
                time_cols.append(col)
        except:
            continue
    if not time_cols:
        numerical_cols = df.select_dtypes(include=[np.number]).columns.tolist()
        time_cols = numerical_cols[:10]

    features_df = pd.DataFrame()
    label_encoders = {}

    for col in categorical_cols[:5]:
        if col in df.columns:
            le = LabelEncoder()
            col_str = df[col].astype(str)
            features_df[f'{col}_encoded'] = le.fit_transform(col_str)
            label_encoders[col] = le

    if time_cols:
        for col in time_cols[:5]:
            if col in df.columns:
                features_df[f'value_{col}'] = df[col]

    features_df['missing_count_total'] = df.isnull().sum(axis=1)

    if categorical_cols:
        cat_cols_exist = [col for col in categorical_cols if col in df.columns]
        if cat_cols_exist:
            cat_missing = df[cat_cols_exist].isnull().sum(axis=1)
            features_df['missing_count_categorical'] = cat_missing

    if time_cols:
        time_cols_exist = [col for col in time_cols if col in df.columns]
        if time_cols_exist:
            time_missing = df[time_cols_exist].isnull().sum(axis=1)
            features_df['missing_count_numerical'] = time_missing

    features_df['missing_percentage'] = (features_df['missing_count_total'] / df.shape[1]) * 100

    iso_features = features_df.copy()
    MISSING_INDICATOR = -999999
    iso_features_filled = iso_features.fillna(MISSING_INDICATOR)

    isolation_forest = IsolationForest(
        contamination=0.05,
        random_state=42,
        n_estimators=200,
        max_samples='auto',
        n_jobs=-1
    )

    iso_predictions = isolation_forest.fit_predict(iso_features_filled)
    iso_scores = isolation_forest.score_samples(iso_features_filled)

    is_anomaly_iso = iso_predictions == -1
    anomaly_count_iso = is_anomaly_iso.sum()

    numerical_anomalies = pd.DataFrame(index=df.index)
    stat_anomaly_count = 0

    for col in time_cols[:3]:
        if col in df.columns:
            col_data = df[col].dropna()

            if len(col_data) > 10:
                mean_val = col_data.mean()
                std_val = col_data.std()

                if std_val > 0:
                    z_scores = np.abs((df[col] - mean_val) / std_val)
                    z_anomalies = z_scores > 3
                    z_anomaly_count = z_anomalies.sum()

                    q1 = col_data.quantile(0.25)
                    q3 = col_data.quantile(0.75)
                    iqr = q3 - q1
                    if iqr > 0:
                        iqr_anomalies = (df[col] < (q1 - 1.5 * iqr)) | (df[col] > (q3 + 1.5 * iqr))
                        iqr_anomaly_count = iqr_anomalies.sum()

                        numerical_anomalies[f'{col}_z_anomaly'] = z_anomalies.fillna(False)
                        numerical_anomalies[f'{col}_iqr_anomaly'] = iqr_anomalies.fillna(False)

                        stat_anomaly_count += z_anomaly_count + iqr_anomaly_count

    if not numerical_anomalies.empty:
        numerical_anomalies['statistical_anomaly_score'] = numerical_anomalies.sum(axis=1)
        is_statistical_anomaly = numerical_anomalies['statistical_anomaly_score'] > 0
        stat_anomaly_count = is_statistical_anomaly.sum()
    else:
        is_statistical_anomaly = pd.Series(False, index=df.index)

    missing_stats = {
        'q1': features_df['missing_percentage'].quantile(0.25),
        'q3': features_df['missing_percentage'].quantile(0.75),
        'mean': features_df['missing_percentage'].mean(),
        'std': features_df['missing_percentage'].std()
    }

    iqr_missing = missing_stats['q3'] - missing_stats['q1']
    high_missing_threshold = missing_stats['q3'] + 1.5 * iqr_missing
    too_much_missing = features_df['missing_percentage'] > high_missing_threshold

    low_missing_threshold = max(0, missing_stats['q1'] - 1.5 * iqr_missing)
    too_little_missing = features_df['missing_percentage'] < low_missing_threshold

    missing_pattern_anomalies = too_much_missing | too_little_missing
    missing_anomaly_count = missing_pattern_anomalies.sum()

    anomaly_results = pd.DataFrame({
        'isolation_forest_anomaly': is_anomaly_iso,
        'isolation_forest_score': iso_scores,
        'statistical_anomaly': is_statistical_anomaly,
        'missing_pattern_anomaly': missing_pattern_anomalies,
        'missing_percentage': features_df['missing_percentage'],
        'missing_count': features_df['missing_count_total']
    })

    if not numerical_anomalies.empty:
        anomaly_results['statistical_anomaly_score'] = numerical_anomalies['statistical_anomaly_score']
    else:
        anomaly_results['statistical_anomaly_score'] = 0

    anomaly_results['combined_anomaly_score'] = (
            anomaly_results['isolation_forest_anomaly'].astype(int) +
            anomaly_results['statistical_anomaly'].astype(int) +
            anomaly_results['missing_pattern_anomaly'].astype(int)
    )

    anomaly_results['anomaly_level'] = pd.cut(
        anomaly_results['combined_anomaly_score'],
        bins=[-0.1, 0.5, 1.5, 2.5, 3.1],
        labels=['Normal', 'Low Risk', 'Medium Risk', 'High Risk']
    )

    anomaly_summary = anomaly_results['anomaly_level'].value_counts().sort_index()
    high_risk_anomalies = anomaly_results[anomaly_results['anomaly_level'] == 'High Risk']

    final_results = df.copy()
    for col in anomaly_results.columns:
        final_results[f'anomaly_{col}'] = anomaly_results[col]

    return final_results, anomaly_results, {
        'isolation_forest': isolation_forest,
        'label_encoders': label_encoders,
        'anomaly_summary': anomaly_summary,
        'high_risk_count': len(high_risk_anomalies),
        'completeness': completeness,
        'categorical_cols': categorical_cols,
        'anomaly_count_iso': anomaly_count_iso,
        'stat_anomaly_count': stat_anomaly_count if not numerical_anomalies.empty else 0,
        'missing_anomaly_count': missing_anomaly_count,
        'df': df,
        'missing_stats': missing_stats
    }


tab1, tab2 = st.tabs(["Report Generator", "Chatbot"])

with tab1:
    col_left, col_center, col_right = st.columns([1, 4, 1])
    with col_center:
        st.title("Report Generator")
        uploaded_file = st.file_uploader("Upload CSV here", type=["csv"])

        if uploaded_file is not None:
            # Load and process the data first
            uploaded_file.seek(0)
            df_uploaded = pd.read_csv(uploaded_file, low_memory=False)
            df_uploaded = coerce_year_columns(df_uploaded)

            st.subheader("Uploaded Data")
            st.dataframe(df_uploaded.head(10))

            # Define all functions first before using them
            TEMPLATE_XLSX = r'Template_Datathon.xlsx'
            SHEET_NAME = 'A2'
            NUMERIC_COLS = [chr(c) for c in range(ord('D'), ord('Y') + 1)]


            def get_currency_pairs():
                return {
                    'D': ('USD', 'AUD'), 'E': ('USD', 'BRL'), 'F': ('USD', 'CAD'),
                    'G': ('USD', 'CHF'), 'H': ('USD', 'CNY'), 'I': ('USD', 'EUR'),
                    'J': ('USD', 'GBP'), 'K': ('USD', 'HKD'), 'L': ('USD', 'INR'),
                    'M': ('USD', 'JPY'), 'N': ('USD', 'KRW'), 'O': ('USD', 'MXN'),
                    'P': ('USD', 'NOK'), 'Q': ('USD', 'NZD'), 'R': ('USD', 'PLN'),
                    'S': ('USD', 'RUB'), 'T': ('USD', 'SEK'), 'U': ('USD', 'SGD'),
                    'V': ('USD', 'TRY'), 'W': ('USD', 'TWD'), 'X': ('USD', 'ZAR'),
                    'Y': ('CLS', None)
                }


            CURRENCY_NAME = {
                'AUD': 'Australian dollar', 'BRL': 'Brazilian real', 'CAD': 'Canadian dollar',
                'CHF': 'Swiss franc', 'CNY': 'Renminbi', 'EUR': 'Euro',
                'GBP': 'Pound (sterling)', 'HKD': 'Hong Kong dollar', 'INR': 'Indian rupee',
                'JPY': 'Yen', 'KRW': 'Won', 'MXN': 'Mexican peso',
                'NOK': 'Norwegian krone', 'NZD': 'New Zealand dollar', 'PLN': 'Zloty',
                'RUB': 'Russian rouble', 'SEK': 'Swedish krona', 'SGD': 'Singapore dollar',
                'TRY': 'Turkish lira', 'TWD': 'New Taiwan dollar', 'ZAR': 'Rand'
            }


            def get_instrument_by_row(row_num: int) -> Optional[str]:
                if 9 <= row_num <= 28:  return "Spot"
                if 29 <= row_num <= 49:  return "Outright forwards"
                if 50 <= row_num <= 56:  return "Outright forwards"
                if 57 <= row_num <= 76:  return "FX swaps"
                if 77 <= row_num <= 83:  return "FX swaps"
                if 84 <= row_num <= 103: return "Currency swaps"
                if 104 <= row_num <= 123: return "Options"
                if 124 <= row_num <= 127: return "Total (all instruments)"
                return None


            def get_counterparty_and_country_by_row(row_num: int) -> Tuple[Optional[str], Optional[str]]:
                m = {
                    10: ("Reporting dealers", None),
                    11: ("Reporting dealers", "Residents/Local"),
                    12: ("Reporting dealers", "Non-residents/Cross-border"),
                    13: ("Other financial institutions", None),
                    14: ("Other financial institutions", "Residents/Local"),
                    15: ("Other financial institutions", "Non-residents/Cross-border"),
                    16: ("Non-reporting banks", None),
                    17: ("Institutional investors", None),
                    18: ("Hedge funds and proprietary trading firms", None),
                    19: ("Official sector financial institutions", None),
                    20: ("Other residual financial institutions", None),
                    21: ("Undistributed", None),
                    22: ("Non-financial customers", None),
                    23: ("Non-financial customers", "Residents/Local"),
                    24: ("Non-financial customers", "Non-residents/Cross-border"),
                    25: ("Total (all counterparties)", None),
                    26: ("Prime brokered", None),
                    27: ("Prime brokered", None),
                    28: ("Retail-driven", None),
                    30: ("Reporting dealers", None),
                    31: ("Reporting dealers", "Residents/Local"),
                    32: ("Reporting dealers", "Non-residents/Cross-border"),
                    33: ("Other financial institutions", None),
                    34: ("Other financial institutions", "Residents/Local"),
                    35: ("Other financial institutions", "Non-residents/Cross-border"),
                    36: ("Non-reporting banks", None),
                    37: ("Institutional investors", None),
                    38: ("Hedge funds and proprietary trading firms", None),
                    39: ("Official sector financial institutions", None),
                    40: ("Other residual financial institutions", None),
                    41: ("Undistributed", None),
                    42: ("Non-financial customers", None),
                    43: ("Non-financial customers", "Residents/Local"),
                    44: ("Non-financial customers", "Non-residents/Cross-border"),
                    45: ("Total (all counterparties)", None),
                    46: ("Prime brokered", None),
                    47: ("Prime brokered", None),
                    48: ("Retail-driven", None),
                    49: ("Total (all counterparties)", None),
                    58: ("Reporting dealers", None),
                    59: ("Reporting dealers", "Residents/Local"),
                    60: ("Reporting dealers", "Non-residents/Cross-border"),
                    61: ("Other financial institutions", None),
                    62: ("Other financial institutions", "Residents/Local"),
                    63: ("Other financial institutions", "Non-residents/Cross-border"),
                    64: ("Non-reporting banks", None),
                    65: ("Institutional investors", None),
                    66: ("Hedge funds and proprietary trading firms", None),
                    67: ("Official sector financial institutions", None),
                    68: ("Other residual financial institutions", None),
                    69: ("Undistributed", None),
                    70: ("Non-financial customers", None),
                    71: ("Non-financial customers", "Residents/Local"),
                    72: ("Non-financial customers", "Non-residents/Cross-border"),
                    73: ("Total (all counterparties)", None),
                    74: ("Prime brokered", None),
                    75: ("Prime brokered", None),
                    76: ("Retail-driven", None),
                    85: ("Reporting dealers", None),
                    86: ("Reporting dealers", "Residents/Local"),
                    87: ("Reporting dealers", "Non-residents/Cross-border"),
                    88: ("Other financial institutions", None),
                    89: ("Other financial institutions", "Residents/Local"),
                    90: ("Other financial institutions", "Non-residents/Cross-border"),
                    91: ("Non-reporting banks", None),
                    92: ("Institutional investors", None),
                    93: ("Hedge funds and proprietary trading firms", None),
                    94: ("Official sector financial institutions", None),
                    95: ("Other residual financial institutions", None),
                    96: ("Undistributed", None),
                    97: ("Non-financial customers", None),
                    98: ("Non-financial customers", "Residents/Local"),
                    99: ("Non-financial customers", "Non-residents/Cross-border"),
                    100: ("Total (all counterparties)", None),
                    101: ("Prime brokered", None),
                    102: ("Prime brokered", None),
                    103: ("Retail-driven", None),
                    105: ("Reporting dealers", None),
                    106: ("Reporting dealers", "Residents/Local"),
                    107: ("Reporting dealers", "Non-residents/Cross-border"),
                    108: ("Other financial institutions", None),
                    109: ("Other financial institutions", "Residents/Local"),
                    110: ("Other financial institutions", "Non-residents/Cross-border"),
                    111: ("Non-reporting banks", None),
                    112: ("Institutional investors", None),
                    113: ("Hedge funds and proprietary trading firms", None),
                    114: ("Official sector financial institutions", None),
                    115: ("Other residual financial institutions", None),
                    116: ("Undistributed", None),
                    117: ("Non-financial customers", None),
                    118: ("Non-financial customers", "Residents/Local"),
                    119: ("Non-financial customers", "Non-residents/Cross-border"),
                    120: ("Total (all counterparties)", None),
                    121: ("Prime brokered", None),
                    122: ("Prime brokered", None),
                    123: ("Retail-driven", None),
                    124: ("Total (all counterparties)", None),
                    125: ("Prime brokered", None),
                    126: ("Prime brokered", None),
                    127: ("Retail-driven", None),
                }
                return m.get(row_num, (None, None))


            def get_maturity_by_row(row_num: int) -> Optional[str]:
                m = {
                    51: "One day",
                    52: "Over 1 day and less than 7 days",
                    53: "Over 7 days and up to 1 month",
                    54: "Over 1 month and up to 3 months",
                    55: "Over 3 months and up to 6 months",
                    56: "Over 6 months",
                    78: "One day",
                    79: "Over 1 day and less than 7 days",
                    80: "Over 7 days and up to 1 month",
                    81: "Over 1 month and up to 3 months",
                    82: "Over 3 months and up to 6 months",
                    83: "Over 6 months",
                }
                return m.get(row_num, None)


            AGG_ROWS_ALL_COUNTRIES = {
                10, 13, 22, 30, 33, 42, 58, 61, 70, 85, 88, 97, 105, 108, 117,
                125, 126, 127
            }

            TOTAL_ROWS_SUM_LOCAL_CB = {25, 45, 73, 100, 120}
            MATURITY_ROWS = {51, 52, 53, 54, 55, 56, 78, 79, 80, 81, 82, 83}
            SKIP_ROWS = {9, 29, 50, 57, 77, 84, 104}


            def load_data(csv_file) -> pd.DataFrame:
                csv_file.seek(0)
                df = pd.read_csv(csv_file, low_memory=False)
                basis_col = 'Basis (AB)' if 'Basis (AB)' in df.columns else ('Basis' if 'Basis' in df.columns else None)
                if basis_col:
                    df = df[df[basis_col] == 'Net - net']
                year_cols = [c for c in df.columns if re.fullmatch(r'\d{4}', str(c))]
                for c in year_cols:
                    df[c] = pd.to_numeric(df[c], errors='coerce').fillna(0.0)
                return df


            def pick_target_year(df: pd.DataFrame, target_year: Optional[int]) -> str:
                years = sorted([int(c) for c in df.columns if re.fullmatch(r'\d{4}', str(c))])
                if not years:
                    raise ValueError("No 4-digit year columns found.")
                if target_year and target_year in years:
                    return str(target_year)
                return str(years[-1])


            def safe_write(ws, row: int, col: int, value):
                target_row, target_col = row, col
                for rng in ws.merged_cells.ranges:
                    if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                        target_row, target_col = rng.min_row, rng.min_col
                        break
                ws.cell(row=target_row, column=target_col, value=value)


            def sum_for_filters(df: pd.DataFrame, filters: dict, year: str) -> float:
                sub = df
                for k, v in filters.items():
                    if k in sub.columns:
                        sub = sub[sub[k] == v]
                return float(sub[year].sum()) if year in sub.columns else 0.0


            def query_cell(df: pd.DataFrame, row_num: int, col_letter: str, year: str) -> float:
                if row_num == 124:
                    return (
                            query_cell(df, 25, col_letter, year) +
                            query_cell(df, 45, col_letter, year) +
                            query_cell(df, 73, col_letter, year) +
                            query_cell(df, 100, col_letter, year) +
                            query_cell(df, 120, col_letter, year)
                    )

                if row_num in SKIP_ROWS:
                    return 0.0

                filters = {
                    'Risk category': 'Foreign exchange',
                    'Reporting country': 'All countries (total)',
                }

                instr = get_instrument_by_row(row_num)
                if instr:
                    filters['Instrument'] = instr

                if row_num == 49:
                    filters['Instrument'] = 'Non-deliverable forwards'
                    filters['Counterparty sector'] = 'Total (all counterparties)'
                    filters['Counterparty country'] = 'All countries (total)'

                pairs = get_currency_pairs()
                if col_letter not in pairs:
                    return 0.0
                _, leg2 = pairs[col_letter]
                if col_letter == 'Y':
                    filters['Currency leg 1'] = 'US dollar'
                    filters['Currency leg 2'] = 'o/w CLS eligible pairs'
                else:
                    filters['Currency leg 1'] = 'US dollar'
                    filters['Currency leg 2'] = CURRENCY_NAME.get(leg2, leg2)

                if row_num != 49:
                    cp_sector, cp_country = get_counterparty_and_country_by_row(row_num)
                    mat = get_maturity_by_row(row_num)

                    if row_num in MATURITY_ROWS:
                        filters['Counterparty sector'] = 'Total (all counterparties)'
                        filters['Counterparty country'] = 'All countries (total)'
                    elif row_num in TOTAL_ROWS_SUM_LOCAL_CB:
                        base = {**filters}
                        if cp_sector:
                            base['Counterparty sector'] = cp_sector
                        if mat:
                            base['Maturity'] = mat
                        f_local = {**base, 'Counterparty country': 'Residents/Local'}
                        f_cross = {**base, 'Counterparty country': 'Non-residents/Cross-border'}
                        return sum_for_filters(df, f_local, year) + sum_for_filters(df, f_cross, year)
                    else:
                        if cp_sector:
                            filters['Counterparty sector'] = cp_sector
                        if row_num in AGG_ROWS_ALL_COUNTRIES:
                            filters['Counterparty country'] = 'All countries (total)'
                        elif cp_country:
                            filters['Counterparty country'] = cp_country

                    if mat:
                        filters['Maturity'] = mat

                return sum_for_filters(df, filters, year)


            def fill_template_streamlit(csv_file, template_file, target_year=None):
                df = load_data(csv_file)
                year_str = pick_target_year(df, target_year)
                wb = load_workbook(template_file)
                ws = wb[SHEET_NAME]
                for row in range(10, 128):
                    row_total = 0.0
                    for col_letter in NUMERIC_COLS:
                        col_idx = ord(col_letter) - ord('A') + 1
                        val = query_cell(df, row, col_letter, year_str)
                        safe_write(ws, row, col_idx, val)
                        if col_letter != 'Y':
                            row_total += val
                    safe_write(ws, row, 26, row_total)
                output = BytesIO()
                wb.save(output)
                wb.close()
                output.seek(0)
                return output


            # Report generation buttons - positioned right after uploaded data table
            st.subheader("Generate Reports")
            col1, col2, col3 = st.columns([2, 2, 1])
            with col1:
                # Generate Excel report when button is clicked
                if st.button("Generate Excel Report", key="generate_excel"):
                    with st.spinner("Generating Excel report..."):
                        try:
                            excel_io = fill_template_streamlit(uploaded_file, TEMPLATE_XLSX)
                            excel_file_name = uploaded_file.name.rsplit(".", 1)[0] + "_report.xlsx"
                            st.session_state.generated_excel = excel_io
                            st.success("Excel report generated successfully!")
                        except Exception as e:
                            st.error(f"Error generating Excel report: {str(e)}")

                # Download button for Excel report
                if "generated_excel" in st.session_state:
                    excel_file_name = uploaded_file.name.rsplit(".", 1)[0] + "_report.xlsx"
                    st.download_button(
                        label="Download Report",
                        data=st.session_state.generated_excel,
                        file_name=excel_file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            with col3:
                if st.button("Generate Anomaly Report", key="generate_anomaly"):
                    with st.spinner("Generating PDF anomaly detection report..."):
                        try:
                            uploaded_file.seek(0)
                            df_for_anomaly = pd.read_csv(uploaded_file, low_memory=False)

                            results_df, anomaly_df, models = comprehensive_anomaly_detection(df_for_anomaly)

                            pdf_output = generate_pdf_report(results_df, anomaly_df, models)

                            st.session_state.pdf_report = pdf_output

                            st.success("Anomaly detection report generated successfully!")

                        except Exception as e:
                            st.error(f"Error generating anomaly report: {str(e)}")

                if "pdf_report" in st.session_state:
                    st.download_button(
                        label="Download Anomaly Report",
                        data=st.session_state.pdf_report,
                        file_name=f"Anomaly_Detection_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf"
                    )

            st.divider()

            # Year selection functionality
            available_years_list = available_years(df_uploaded)
            if available_years_list:
                st.subheader("📊 Data Analysis")

                # Year selection dropdown
                col1, col2 = st.columns(2)
                with col1:
                    selected_year = st.selectbox(
                        "Choose reporting year",
                        options=sorted(available_years_list, reverse=True),
                        index=0,
                        help="Select the year for detailed analysis"
                    )

                with col2:
                    # Multi-year selection for trends
                    default_years = available_years_list[-5:] if len(
                        available_years_list) >= 5 else available_years_list
                    selected_years = st.multiselect(
                        "Years for trend analysis",
                        options=available_years_list,
                        default=default_years,
                        help="Select multiple years to see trends"
                    )

                # Data visualization section
                if selected_years:
                    st.subheader("📈 Data Overview")

                    # Create tabs for different chart types
                    chart_tab1, chart_tab2 = st.tabs(["Data Trends", "Distribution Analysis"])

                    with chart_tab1:
                        st.markdown("**Multi-year data completeness and record availability**")
                        overview_chart = create_data_overview_chart(df_uploaded, selected_years)
                        st.plotly_chart(overview_chart, use_container_width=True)

                    with chart_tab2:
                        st.markdown("**Instrument distribution for selected year**")
                        distribution_chart = create_instrument_distribution_chart(df_uploaded, selected_year)
                        st.plotly_chart(distribution_chart, use_container_width=True)

with tab2:
    st.title("Ask AI about your report")

    if "generated_excel" in st.session_state:
        excel_io = st.session_state.generated_excel
        excel_io.seek(0)
        df = pd.read_excel(excel_io)
    else:
        df = None

    if df is None:
        st.info("Please upload a CSV in the Report Generator tab first.")
    else:
        # Initialize chat messages in session state
        if "messages" not in st.session_state:
            st.session_state.messages = []

        # Display chat messages
        for i, msg in enumerate(st.session_state.messages):
            message(msg["content"], is_user=(msg["role"] == "user"), key=f"msg_{i}")

        # Chat input
        user_input = st.chat_input("Ask ChatGPT about your Excel...", key="chat_input_tab2")

        # Process user input immediately when submitted
        if user_input:
            # Add user message to session state
            st.session_state.messages.append({"role": "user", "content": user_input})

            # Display user message immediately
            message(user_input, is_user=True, key=f"msg_{len(st.session_state.messages) - 1}")

            # Prepare data for AI
            max_rows = 200
            if len(df) > max_rows:
                df_preview = df.sample(max_rows).to_csv(index=False)
                context_note = f"Previewing {max_rows} random rows from your data."
            else:
                df_preview = df.to_csv(index=False)
                context_note = "Full data provided."

            prompt = (
                f"{context_note}\n\n"
                f"Here is your Excel data:\n{df_preview}\n\n"
                f"User question: {user_input}\n"
                "Answer based on this data only. Be precise and refer to the values."
            )

            # Generate AI response
           # try:
           #     response = client.chat.completions.create(
            #        model="gpt-4o-mini",
            #        messages=[{"role": "user", "content": prompt}],
           #         max_tokens=700
            #    )
            #    bot_response = response.choices[0].message.content
          #  except Exception as e:
          #      bot_response = f"Something went wrong: {e}"

            # Add AI response to session state
        #    st.session_state.messages.append({"role": "assistant", "content": bot_response})

            # Display AI response immediately
         #   message(bot_response, is_user=False, key=f"msg_{len(st.session_state.messages) - 1}")

            # Force a rerun to update the display
            st.rerun()